
"""
MAC Address OCR Scanner — Debugged & Optimized v2.1
====================================================
Root-cause analysis of original failures:
  1. Full-image OCR: The LCD screen is too small relative to the whole photo.
     EasyOCR gets confused by keypad labels, boxes, etc.
  2. Wrong allowlist: Using hex-only allowlist blocked EasyOCR's internal
     language model from correcting OCR errors contextually.
  3. Threshold 92% on 12 chars = only 1 error allowed. LCD photos easily
     produce 2-3 character misreads.
  4. No OCR character corrections (O→0, S→5, Z→2, etc.)
  5. Hex-group extraction via strict word boundaries missed groups
     embedded in "BC AB;F5 A2C5:35" style output.

Fixed strategy:
  1. Crop 3 candidate screen regions from the image (different phone orientations)
  2. Try all 4 rotations on each crop
  3. Extract MAC via flexible regex that tolerates OCR noise between octets
  4. Apply OCR character corrections before matching
  5. Fall back to fuzzy sliding-window on full cleaned hex string
"""

import os
import re
import shutil
import warnings
import numpy as np
from pathlib import Path
from PIL import Image, ImageEnhance, ImageFilter
import cv2
import easyocr
from thefuzz import fuzz

warnings.filterwarnings("ignore", category=UserWarning)


# ─────────────────────────────────────────────
# OCR Character Corrections
# ─────────────────────────────────────────────
# Common substitutions for small LCD text OCR errors.
# Only safe replacements — letters that are never valid hex digits.
OCR_CORRECTIONS = str.maketrans({
    'O': '0',  # Letter O → digit 0
    'I': '1',  # Letter I → digit 1
    'L': '1',  # Letter L → digit 1
    'S': '5',  # Letter S → digit 5
    'Z': '2',  # Letter Z → digit 2
    'Q': '0',  # Letter Q → digit 0
    # NOTE: G→6 skipped: G is not a hex digit but 6 looks like G in some fonts
    # NOTE: B→8 skipped: B is a valid hex digit!
    # NOTE: D→0 skipped: D is a valid hex digit!
})


def apply_corrections(text: str) -> str:
    """Apply OCR character corrections to uppercase text."""
    return text.upper().translate(OCR_CORRECTIONS)


# ─────────────────────────────────────────────
# MAC Input Validation
# ─────────────────────────────────────────────
def clean_mac_target(mac_input: str) -> str:
    """Strip separators and validate a MAC address string."""
    cleaned = re.sub(r'[\s:.\-]', '', mac_input).upper()
    if not re.fullmatch(r'[0-9A-F]{12}', cleaned):
        raise ValueError(f"Invalid MAC format: '{mac_input}'")
    return cleaned


# ─────────────────────────────────────────────
# MAC Extraction from OCR Text
# ─────────────────────────────────────────────
def extract_macs_from_text(raw_text: str) -> list[str]:
    """
    Extract 12-char hex MAC addresses from OCR text.

    Primary strategy: Look for text AFTER the 'MAC' keyword — anchors extraction
    so we don't accidentally include hex chars from preceding words like
    'ADDRE55' → '55' being mistaken for the first octet.

    Fallback strategy: Flexible 6-octet regex (may produce false positives,
    checked by fuzzy comparison against known target prefixes).
    """
    corrected = apply_corrections(raw_text)
    macs = []

    # ── Strategy A: Anchor extraction after 'MAC' keyword (most accurate) ──
    # Handles: "MAC Address: BC AB F5 A2,C5. 35"
    #          "MAC Address; BC AB;FS A2C5:35"
    #          "MAC Addre55: BC ABHF6 A2C4 DF"
    mac_kw = re.search(r'MAC\s*(?:ADDRE?\w*)?\s*[:;,]?\s*', corrected)
    if mac_kw:
        snippet = corrected[mac_kw.end():mac_kw.end() + 60]
        # Extract all 2-char hex groups from the snippet
        hex_pairs = re.findall(r'[0-9A-F]{2}', snippet)
        if len(hex_pairs) >= 6:
            candidate = ''.join(hex_pairs[:6])
            macs.append(candidate)
            # Also try starting from pair index 1 in case first pair is noise
            if len(hex_pairs) >= 7:
                macs.append(''.join(hex_pairs[1:7]))

    # ── Strategy B: Flexible 6-octet regex over full text (fallback) ──
    # Allows 0-3 separator chars between each octet pair.
    mac_pattern = re.compile(
        r'([0-9A-F]{2})[^0-9A-F]{0,3}'
        r'([0-9A-F]{2})[^0-9A-F]{0,3}'
        r'([0-9A-F]{2})[^0-9A-F]{0,3}'
        r'([0-9A-F]{2})[^0-9A-F]{0,3}'
        r'([0-9A-F]{2})[^0-9A-F]{0,3}'
        r'([0-9A-F]{2})'
    )
    for match in mac_pattern.finditer(corrected):
        candidate = ''.join(match.groups())
        if candidate not in macs:
            macs.append(candidate)

    return macs


def char_match_score(a: str, b: str) -> int:
    """
    Hamming-style character match score (0-100) for equal-length strings.

    WHY NOT fuzz.ratio:
    fuzz.ratio uses SequenceMatcher which finds common *substrings* and
    therefore scores reordered/shifted characters higher than they deserve.
    Example: 'BCABF5A2C9C6' vs 'BCABF5A2C697' — SequenceMatcher finds
    'BCABF5A2C' (9) + '6' (1) = 10 common chars → 83% score.
    But position-by-position only 9/12 match → 75% Hamming.
    Since ALL target MACs share the prefix 'BCABF5A2C', SequenceMatcher
    cannot reliably distinguish between them. Hamming is correct here.
    """
    if len(a) != len(b) or not a:
        return 0
    matches = sum(1 for x, y in zip(a, b) if x == y)
    return int(100 * matches / len(a))


def fuzzy_find_in_hex(target_mac: str, clean_hex_str: str, threshold: int = 83) -> tuple[bool, int]:
    """
    Sliding-window Hamming match on a stripped hex string.
    Fallback for when structured extraction fails.
    threshold=83 → 10/12 chars must match (≤2 errors allowed).
    """
    tlen = len(target_mac)
    slen = len(clean_hex_str)
    if slen < tlen:
        return False, 0

    best = 0
    for i in range(slen - tlen + 1):
        window = clean_hex_str[i:i + tlen]
        score = char_match_score(target_mac, window)
        if score > best:
            best = score
        if score >= threshold:
            return True, score
    return False, best


# ─────────────────────────────────────────────
# Image Preprocessing
# ─────────────────────────────────────────────
def enhance_for_ocr(pil_img: Image.Image) -> list[np.ndarray]:
    """
    Return multiple preprocessed variants of the image.
    Multiple variants increase OCR hit rate for different lighting conditions.
    """
    variants = []
    img_rgb = np.array(pil_img.convert('RGB'))

    # Variant 1: Original
    variants.append(img_rgb)

    # Variant 2: CLAHE (adaptive contrast) — best for LCD with ambient light
    gray = cv2.cvtColor(img_rgb, cv2.COLOR_RGB2GRAY)
    clahe = cv2.createCLAHE(clipLimit=3.0, tileGridSize=(8, 8))
    gray_eq = clahe.apply(gray)
    variants.append(cv2.cvtColor(gray_eq, cv2.COLOR_GRAY2RGB))

    # Variant 3: Sharpened + high contrast
    pil_sharp = ImageEnhance.Contrast(pil_img.convert('RGB')).enhance(2.0)
    pil_sharp = ImageEnhance.Sharpness(pil_sharp).enhance(3.0)
    pil_sharp = pil_sharp.filter(ImageFilter.UnsharpMask(radius=2, percent=200, threshold=3))
    variants.append(np.array(pil_sharp))

    return variants


# ─────────────────────────────────────────────
# Screen Region Crop Candidates
# ─────────────────────────────────────────────
def get_screen_crop_candidates(img: Image.Image) -> list[Image.Image]:
    """
    Return candidate crops targeting the phone LCD screen region.
    
    From observation: The Cisco IP Phone 7800 screen is in the upper-right
    quadrant of the image (the phone is photographed from above/front).
    We try multiple crop ratios to handle different photo compositions.
    """
    w, h = img.size
    crops = []

    # Primary crop: right-center area (most consistent placement)
    crops.append(img.crop((int(w * 0.45), int(h * 0.30), int(w * 0.92), int(h * 0.78))))

    # Wider crop: in case phone is further away or differently framed
    crops.append(img.crop((int(w * 0.35), int(h * 0.25), int(w * 0.95), int(h * 0.85))))

    # Full image (fallback): in case the phone is oriented differently
    crops.append(img)

    return crops


# ─────────────────────────────────────────────
# Single Image Scanner
# ─────────────────────────────────────────────
def scan_image(image_path: Path, targets: dict, reader: easyocr.Reader, debug: bool = False) -> list:
    """
    Scan a single image for target MAC addresses.
    Returns list of clean MAC strings that were matched.
    """
    found = []
    try:
        img = Image.open(image_path)
        # Preserve resolution for small text — do NOT shrink below original if smaller
        img.thumbnail((3000, 3000), Image.Resampling.LANCZOS)

        # Get crop candidates centered on the phone screen
        crop_candidates = get_screen_crop_candidates(img)

        for crop_idx, crop in enumerate(crop_candidates):
            for angle in [90, 0, 270, 180]:  # 90° first: most common phone orientation
                rotated = crop.rotate(angle, expand=True)
                img_variants = enhance_for_ocr(rotated)

                # Collect all OCR tokens across preprocessing variants
                all_raw_tokens = []
                for variant_np in img_variants:
                    try:
                        # No allowlist — let EasyOCR use its full language model
                        results = reader.readtext(variant_np, detail=0, paragraph=False)
                        all_raw_tokens.extend(results)
                    except Exception:
                        pass

                if not all_raw_tokens:
                    continue

                full_text = " ".join(all_raw_tokens)

                if debug:
                    print(f"\n  [DEBUG] crop={crop_idx} angle={angle}°:")
                    print(f"    raw: {full_text[:120]}")

                # ── Strategy 1: Structured MAC extraction (most accurate) ──
                extracted_macs = extract_macs_from_text(full_text)

                if debug and extracted_macs:
                    print(f"    extracted MACs: {extracted_macs}")

                # ── Strategy 2: Fuzzy match on cleaned hex string (fallback) ──
                corrected_text = apply_corrections(full_text)
                clean_hex = re.sub(r'[^0-9A-F]', '', corrected_text)

                angle_matched = False

                for clean_mac, info in targets.items():
                    if info["found_in"] is not None:
                        continue  # Already found in a previous image
                    if clean_mac in found:
                        continue  # Already found in this image

                    # Check strategy 1: exact match in extracted MACs
                    matched = False
                    match_method = ""
                    best_score = 0

                    for extracted in extracted_macs:
                        # Slide over the extracted MAC to handle off-by-one prefix pollution
                        ex_len = len(extracted)
                        for start in range(max(0, ex_len - 12), ex_len - 11):
                            window = extracted[start:start + 12]
                            if len(window) < 12:
                                continue
                            # Use Hamming (char_match_score) not SequenceMatcher:
                            # threshold 90% = 11/12 chars match (≤1 error)
                            # This is strict but correct since extraction is clean
                            score = char_match_score(clean_mac, window)
                            if score > best_score:
                                best_score = score
                            if score >= 90:
                                matched = True
                                match_method = f"structured:{score}%:{window}"
                                break
                        if matched:
                            break

                    # Strategy 2: fuzzy hex fallback (disabled - causes false positives)
                    # All 5 target MACs share prefix 'BCABF5A2C', leaving only 3 unique chars.
                    # With 2 errors allowed, different target MACs become indistinguishable.
                    # Strategy A (MAC keyword extraction) is reliable enough on its own.
                    if False and not matched:
                        fz_matched, fz_score = fuzzy_find_in_hex(clean_mac, clean_hex, threshold=84)
                        if fz_matched:
                            matched = True
                            match_method = f"fuzzy_hex:{fz_score}%"

                    if matched:
                        found.append(clean_mac)
                        angle_matched = True
                        if debug:
                            print(f"    [OK] MATCH {clean_mac} via {match_method}")

                if angle_matched:
                    # Found at least one MAC in this crop+angle, stop rotating this crop
                    break

            # If all remaining targets are found, stop trying more crops
            remaining = [m for m, i in targets.items() if i["found_in"] is None and m not in found]
            if not remaining:
                break

        return found

    except Exception as e:
        print(f"  ERROR reading {image_path.name}: {e}")
        import traceback
        traceback.print_exc()
        return []


# ─────────────────────────────────────────────
# Main Execution
# ─────────────────────────────────────────────
if __name__ == "__main__":
    # Force UTF-8 output on Windows (avoids CP1252 crash on special chars)
    import sys, io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

    # ── Configuration ──────────────────────────
    import openpyxl

    xlsx_file = r"C:\Users\KIE\Desktop\ocr\SN อุปกรณ์โครงการ 119260207_2205 Mac.xlsx"
    mac_column = 5          # Column E (1-indexed)
    mac_sheet_index = 0     # First sheet

    search_directory      = r""
    destination_directory = r"C"
    output_text_file      = r""

    # Set True to print per-image OCR tokens (useful for threshold tuning)
    DEBUG = False
    # ───────────────────────────────────────────

    # ── Load MAC list from Excel column E ──────
    print(f"Loading MAC list from: {xlsx_file}")
    wb_xl = openpyxl.load_workbook(xlsx_file, data_only=True)
    ws_xl = wb_xl.worksheets[mac_sheet_index]
    print(f"  Sheet : {ws_xl.title}  |  Rows: {ws_xl.max_row}")

    raw_mac_list = []   # list of (raw_mac_str, sn_str, row_no_int)
    for row_idx in range(2, ws_xl.max_row + 1):   # skip header row 1
        val = ws_xl.cell(row=row_idx, column=mac_column).value
        if val is None:
            continue
        cleaned = re.sub(r'[\s:.\-]', '', str(val)).upper()
        if re.fullmatch(r'[0-9A-F]{12}', cleaned):
            sn  = ws_xl.cell(row=row_idx, column=3).value or ''   # col C = S/N
            no  = ws_xl.cell(row=row_idx, column=1).value or ''   # col A = NO
            raw_mac_list.append((str(val).strip(), str(sn).strip(), int(no) if str(no).isdigit() else row_idx - 1))

    print(f"  Valid MACs: {len(raw_mac_list)}")
    print(f"  First: {raw_mac_list[0]}  |  Last: {raw_mac_list[-1]}")
    print()


    try:
        Path(destination_directory).mkdir(parents=True, exist_ok=True)

        targets = {}
        for raw_mac, sn, no in raw_mac_list:
            clean = clean_mac_target(raw_mac)
            targets[clean] = {"raw": raw_mac, "sn": sn, "no": no, "found_in": None}

        print("=" * 60)
        print(" MAC Address OCR Scanner — v2.1 (Excel mode)")
        print("=" * 60)
        print(f"Targets: {len(targets)} MACs from Excel column E")
        print(f"Source : {search_directory}")
        print(f"Dest   : {destination_directory}")
        print()
        print("Loading EasyOCR model (first run downloads ~1.5 GB)...")

        reader = easyocr.Reader(['en'], gpu=True)
        print("EasyOCR ready.\n")

        folder_path = Path(search_directory)
        valid_extensions = {'.jpg', '.jpeg', '.png', '.bmp', '.tiff', '.tif'}
        output_path_resolved = Path(output_text_file).resolve()

        scanned = 0
        total_matched = 0

        for file_path in sorted(folder_path.rglob('*')):
            if not (file_path.is_file() and file_path.suffix.lower() in valid_extensions):
                continue
            if file_path.resolve() == output_path_resolved:
                continue

            scanned += 1
            print(f"[{scanned:03d}] {file_path.name:<40}", end="", flush=True)

            matched_macs = scan_image(file_path, targets, reader, debug=DEBUG)

            if matched_macs:
                total_matched += len(matched_macs)
                print()  # newline
                for mac in matched_macs:
                    targets[mac]["found_in"] = str(file_path.absolute())
                    print(f"       [MATCH] {targets[mac]['raw']}")

                    dest_path = Path(destination_directory) / file_path.name
                    if dest_path.exists():
                        print(f"         (already exists in destination, skipped copy)")
                    else:
                        shutil.copy2(file_path, dest_path)
                        print(f"         Copied → {dest_path}")
            else:
                print("[ no match ]")

            # Early exit
            if all(info["found_in"] is not None for info in targets.values()):
                print("\n>>> All MACs found — stopping early.")
                break

        # ── Summary ──────────────────────────────
        print(f"\n{'=' * 55}")
        print(f" RESULTS: {scanned} images scanned, {total_matched}/{len(targets)} MACs found")
        print(f"{'=' * 55}")

        with open(output_text_file, "w", encoding="utf-8") as f:
            f.write("MAC Address OCR Scan Results\n")
            f.write("=" * 60 + "\n")
            f.write(f"{'NO':<5} {'S/N':<15} {'MAC (ColE)':<14} {'STATUS':<8} PATH\n")
            f.write("-" * 60 + "\n")
            for clean_mac, info in targets.items():
                status = "FOUND" if info["found_in"] else "MISSING"
                path   = info["found_in"] or ""
                line   = f"{info['no']:<5} {info['sn']:<15} {info['raw']:<14} {status:<8} {path}"
                print(f"  {line}")
                f.write(line + "\n")

        print(f"\nResults saved → {output_text_file}")

    except Exception:
        import traceback
        print("\n" + "!" * 55)
        print("CRASH REPORT:")
        traceback.print_exc()
        print("!" * 55)
